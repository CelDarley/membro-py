from app import create_app
from flask import current_app
from flask.cli import with_appcontext
import click
from app.db import db
from app.models import User, Membro
import os
import unicodedata

app = create_app()

@app.cli.command('create-admin')
@click.option('--name', prompt=True)
@click.option('--email', prompt=True)
@click.option('--password', prompt=True, hide_input=True)
@with_appcontext
def create_admin(name, email, password):
	user = User.query.filter_by(email=email.lower()).first()
	if user:
		click.echo('Usuário já existe')
		return
	user = User(name=name, email=email.lower(), role='admin')
	user.set_password(password)
	db.session.add(user)
	db.session.commit()
	click.echo(f'Admin criado: {email}')


def _norm(s: str) -> str:
	if s is None:
		return ''
	s = str(s)
	s = unicodedata.normalize('NFD', s)
	s = ''.join(ch for ch in s if unicodedata.category(ch) != 'Mn')
	return s.strip().lower()


@app.cli.command('import-membros')
@click.argument('path')
@click.option('--truncate', is_flag=True, help='Limpa tabelas antes de importar')
@with_appcontext
def import_membros(path, truncate):
	path = os.path.abspath(path)
	if not os.path.exists(path):
		click.echo(f'Arquivo não encontrado: {path}')
		return
	if truncate:
		try:
			# limpar pivot primeiro
			db.session.execute(db.text('DELETE FROM membro_amigos'))
		except Exception:
			pass
		Membro.query.delete()
		db.session.commit()
	rows = []
	ext = os.path.splitext(path)[1].lower()
	# conjuntos para detectar linha de cabeçalho
	expected = set([
		'membro','mamp','sexo','data de nascimento','telefone celular','email','e-mail','concurso','titularidade','cargo efetivo','comarca lotacao','unidade lotacao','telefone unidade','cargo especial'
	])
	if ext == '.xlsx':
		from openpyxl import load_workbook
		wb = load_workbook(path)
		sheet = wb.active
		# detectar linha de cabeçalho nas primeiras 20 linhas
		header_row = None
		max_scan = min(20, sheet.max_row)
		for r in range(1, max_scan+1):
			vals = [str(c.value or '').strip() for c in next(sheet.iter_rows(min_row=r, max_row=r))[0:]]
			norms = [_norm(v) for v in vals]
			matches = sum(1 for v in norms if v in expected)
			if matches >= 3:
				header_row = r
				headers = vals
				break
		if not header_row:
			# fallback: primeira linha
			headers = [str(c.value or '').strip() for c in next(sheet.iter_rows(min_row=1, max_row=1))[0:]]
			data_iter = sheet.iter_rows(min_row=2, values_only=True)
		else:
			data_iter = sheet.iter_rows(min_row=header_row+1, values_only=True)
		for row in data_iter:
			rows.append(list(row))
	elif ext == '.xls':
		import xlrd
		wb = xlrd.open_workbook(path)
		sheet = wb.sheet_by_index(0)
		header_row = None
		max_scan = min(20, sheet.nrows)
		for r in range(0, max_scan):
			vals = [str(sheet.cell_value(r, col) or '').strip() for col in range(sheet.ncols)]
			norms = [_norm(v) for v in vals]
			matches = sum(1 for v in norms if v in expected)
			if matches >= 3:
				header_row = r
				headers = vals
				break
		if header_row is None:
			headers = [str(sheet.cell_value(0, col) or '').strip() for col in range(sheet.ncols)]
			start = 1
		else:
			start = header_row + 1
		for r in range(start, sheet.nrows):
			rows.append([sheet.cell_value(r, c) for c in range(sheet.ncols)])
	else:
		click.echo('Formato não suportado. Use .xls ou .xlsx')
		return

	# construir mapa de cabeçalhos normalizados
	norm_headers = [_norm(h) for h in headers]
	idx_by_norm = { h:i for i,h in enumerate(norm_headers) }

	# aliases por campo (normalizados)
	aliases = {
		'nome': ['membro','nome'],
		'sexo': ['sexo','genero','genero biologico','genero biologico','genero biológico','genero biologico ','gênero'],
		'concurso': ['concurso','classificacao','classificação'],
		'cargo_efetivo': ['cargo efetivo','cargo_efetivo','cargo atual','cargo'],
		'titularidade': ['titularidade'],
		'email_pessoal': ['email pessoal','e-mail pessoal','email  pessoal','emailpessoal','e mail pessoal','email pessoal '],
		'email_institucional': ['email institucional','e-mail institucional','mail institucional','email inst'],
		'cargo_especial': ['cargo especial'],
		'telefone_unidade': ['telefone unidade','tel unidade','telefone da unidade','telefone trabalho'],
		'telefone_celular': ['telefone celular','celular','telefone movel','telefone móvel'],
		'unidade_lotacao': ['unidade lotacao','unidade lotação','lotacao','lotação','unidade'],
		'comarca_lotacao': ['comarca lotacao','comarca lotação','comarca','cidade'],
		'time_extraprofissionais': ['time de futebol e outros grupos extraprofissionais','time extraprofissionais','grupos extraprofissionais','time de futebol'],
		'quantidade_filhos': ['quantidade de filhos','qtd filhos','qtde filhos','numero de filhos','n filhos'],
		'nomes_filhos': ['nome dos filhos','nomes dos filhos'],
		'estado_origem': ['estado de origem','uf origem','uf'],
		'academico': ['academico','acadêmico'],
		'pretensao_carreira': ['pretensao de movimentacao na carreira','pretensão de movimentação na carreira','pretensao carreira'],
		'carreira_anterior': ['carreira anterior'],
		'lideranca': ['lideranca','liderança'],
		'grupos_identitarios': ['grupos identitarios','grupos identitários','grupo identitarios','grupo identitário'],
		'amigos_ids': ['amigos no mp (ids)','amigos mp (ids)','amigos mp ids','amigos (ids)']
	}

	def find_idx(keys):
		for k in keys:
			i = idx_by_norm.get(_norm(k))
			if i is not None:
				return i
		return None

	map_idx = { field: find_idx(al) for field, al in aliases.items() }

	def get(row, field):
		i = map_idx.get(field)
		if i is None:
			return None
		val = row[i]
		if isinstance(val, str):
			val = val.strip()
		return val if val != '' else None

	created = []
	for r in rows:
		email_p = get(r,'email_pessoal')
		email_i = get(r,'email_institucional')
		m = Membro(
			nome = get(r,'nome'),
			sexo = get(r,'sexo'),
			concurso = str(get(r,'concurso') or '') or None,
			cargo_efetivo = get(r,'cargo_efetivo'),
			titularidade = get(r,'titularidade'),
			email_pessoal = (email_p or email_i),
			cargo_especial = get(r,'cargo_especial'),
			telefone_unidade = str(get(r,'telefone_unidade') or '') or None,
			telefone_celular = str(get(r,'telefone_celular') or '') or None,
			unidade_lotacao = get(r,'unidade_lotacao'),
			comarca_lotacao = get(r,'comarca_lotacao'),
			time_extraprofissionais = get(r,'time_extraprofissionais'),
			quantidade_filhos = (int(float(get(r,'quantidade_filhos'))) if (get(r,'quantidade_filhos') not in (None,'')) else None),
			nomes_filhos = get(r,'nomes_filhos'),
			estado_origem = (str(get(r,'estado_origem') or '')[:2].upper() or None),
			academico = get(r,'academico'),
			pretensao_carreira = get(r,'pretensao_carreira'),
			carreira_anterior = get(r,'carreira_anterior'),
			lideranca = get(r,'lideranca'),
			grupos_identitarios = get(r,'grupos_identitarios'),
		)
		db.session.add(m)
		created.append((m, r))
	db.session.commit()

	# relacionamentos por IDs
	amigos_col = map_idx.get('amigos_ids')
	if amigos_col is not None:
		for m, row in created:
			raw_ids = row[amigos_col]
			if raw_ids is None:
				continue
			if not isinstance(raw_ids, str):
				raw_ids = str(raw_ids)
			raw_ids = raw_ids.strip()
			if not raw_ids:
				continue
			ids = [int(x) for x in filter(None, [s.strip() for s in raw_ids.replace(';',',').replace('|',',').split(',')]) if x.isdigit()]
			if not ids:
				continue
			friends = Membro.query.filter(Membro.id.in_(ids)).all()
			for f in friends:
				if f.id != m.id and f not in m.amigos:
					m.amigos.append(f)
		db.session.commit()

	click.echo(f'Importados: {len(created)} (com relacionamentos quando informados)')


@app.cli.command('seed-demo')
@click.option('--force', is_flag=True, help='Limpa as tabelas antes de inserir exemplos')
@with_appcontext
def seed_demo(force):
	# opcionalmente limpar tabelas
	if force:
		try:
			# limpar pivot primeiro
			db.session.execute(db.text('DELETE FROM membro_amigos'))
		except Exception:
			pass
		Membro.query.delete()
		db.session.commit()
	if Membro.query.count() > 0:
		click.echo('Já existem membros, não será duplicado.')
		return
	m1 = Membro(nome='WILSON PENIN COUTO', sexo='Masculino', concurso='2001', cargo_efetivo='Promotor', email_pessoal='wilson@example.com', comarca_lotacao='BELO HORIZONTE')
	m2 = Membro(nome='WESLEY LEITE VAZ', sexo='Masculino', concurso='2002', cargo_efetivo='Promotor', email_pessoal='wesley@example.com', comarca_lotacao='BELO HORIZONTE')
	m3 = Membro(nome='VANESSA CAMPOLINA REBELLO HORTA', sexo='Feminino', concurso='2003', cargo_efetivo='Promotora', email_pessoal='vanessa@example.com', comarca_lotacao='CONTAGEM')
	db.session.add_all([m1,m2,m3])
	db.session.commit()
	# relacionamentos (amigos no MP)
	m1.amigos.append(m2)
	m1.amigos.append(m3)
	db.session.commit()
	click.echo('Seed de membros concluído com 3 registros e relacionamentos.')

if __name__ == '__main__':
	app.run(host='0.0.0.0', port=8000, debug=True) 